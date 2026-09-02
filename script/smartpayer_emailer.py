#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SmartPayer Email Sender
========================
Automatically composes and sends each generated letter PDF to the
designated recipient based on the client name in the filename.

Filename pattern:   Smartpayer_{Month}_{Year}_{CLIENT_NAME}_letter.pdf
Subject pattern:    Smart Payer {CLIENT NAME} Periode {Month} {Year}

Two send methods (tried in order):
  1. SMTP  — uses credentials from smartpayer_email_config.json
  2. Outlook COM — uses the locally installed Outlook app (Windows only,
                   no credentials needed; Outlook must already be signed in)

Standalone usage:
    python smartpayer_emailer.py --pdf output/Smartpayer_April_2026_AGUS_SETIAWAN_letter.pdf
    python smartpayer_emailer.py --watch output/
    python smartpayer_emailer.py --edit-config
"""

import os
import sys
import re
import json
import time
import smtplib
import argparse
import subprocess
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ── Force UTF-8 output (Windows cp1252 safety) ──────────────────────────────
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try: sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

def _print(*args, **kw):
    try:
        print(*args, **kw)
    except UnicodeEncodeError:
        safe = " ".join(str(a).encode("ascii","replace").decode("ascii") for a in args)
        print(safe, **{k:v for k,v in kw.items() if k!="end"})

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent.resolve()
CONFIG_FILE = SCRIPT_DIR / "smartpayer_email_config.json"

# ── Month helpers ─────────────────────────────────────────────────────────────
MONTHS_EN = ["January","February","March","April","May","June",
             "July","August","September","October","November","December"]
MONTHS_ID = ["Januari","Februari","Maret","April","Mei","Juni",
             "Juli","Agustus","September","Oktober","November","Desember"]

def _month_index(name: str) -> int:
    """Return 1-based month index from English or Indonesian name, or 0 if unknown."""
    name = name.strip().title()
    for i, m in enumerate(MONTHS_EN):
        if m.lower() == name.lower():
            return i + 1
    for i, m in enumerate(MONTHS_ID):
        if m.lower() == name.lower():
            return i + 1
    return 0

def _next_month(month_name: str, year: str) -> tuple:
    """Return (next_month_name_EN, year_str) for the month after month_name."""
    idx = _month_index(month_name)
    if idx == 0:
        return (month_name, year)
    next_idx = idx % 12 + 1          # wraps Dec→Jan
    next_year = int(year) + (1 if idx == 12 else 0)
    return (MONTHS_EN[next_idx - 1], str(next_year))


# =============================================================================
# CONFIG
# =============================================================================

def load_config() -> dict:
    if not CONFIG_FILE.exists():
        _print(f"  [!] Config not found: {CONFIG_FILE}")
        _print(f"      Run --edit-config to create it.")
        return {}
    with open(CONFIG_FILE, encoding="utf-8") as f:
        cfg = json.load(f)
    return _migrate_to_accounts(cfg)

def save_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# =============================================================================
# MULTI-ACCOUNT SUPPORT
# =============================================================================
#
# Config can hold several named sender identities under "accounts", e.g.:
#   {
#     "active_account": "arrd",
#     "accounts": {
#       "arrd":     {"smtp": {...}, "outlook_fallback": {...}},
#       "finance":  {"smtp": {...}, "outlook_fallback": {...}}
#     },
#     "cc_always": [...],
#     "recipients": {...},
#     "email_template": {...}
#   }
#
# Older configs stored a single top-level "smtp"/"outlook_fallback" pair —
# those are auto-migrated into an "accounts" entry named "default" the
# first time the config is loaded, so nothing breaks.

def _migrate_to_accounts(cfg: dict) -> dict:
    """Move legacy top-level smtp/outlook_fallback into accounts['default']."""
    if "accounts" in cfg:
        cfg.setdefault("active_account", next(iter(cfg["accounts"]), "default"))
        return cfg
    if "smtp" in cfg or "outlook_fallback" in cfg:
        cfg["accounts"] = {
            "default": {
                "smtp": cfg.pop("smtp", {}),
                "outlook_fallback": cfg.pop("outlook_fallback", {}),
            }
        }
        cfg["active_account"] = "default"
    return cfg

def get_account(cfg: dict, name: str = None) -> dict:
    """
    Return the {"smtp":..., "outlook_fallback":...} dict for the named
    account, or the active account if name is None. Returns {} if not found.
    """
    accounts = cfg.get("accounts", {})
    key = name or cfg.get("active_account")
    return accounts.get(key, {})

def list_accounts(cfg: dict) -> list:
    return list(cfg.get("accounts", {}).keys())

def set_active_account(cfg: dict, name: str) -> bool:
    if name not in cfg.get("accounts", {}):
        return False
    cfg["active_account"] = name
    return True

def get_recipient(cfg: dict, client_name: str) -> dict:
    """
    Look up TO / CC for a client name.
    Case-insensitive; also tries matching if the config key is a substring
    of the filename client name (handles partial matches like
    'PUTRA NIAGA' matching 'PT. PUTRA NIAGA SEJAHTERA').
    Returns {"to": [...], "cc": [...]} or empty lists.
    """
    recips = cfg.get("recipients", {})
    name_upper = client_name.upper().strip()

    # 1. Exact match (case-insensitive)
    for key, val in recips.items():
        if key.upper().strip() == name_upper:
            return val

    # 2. Config key is contained in filename client name
    for key, val in recips.items():
        if key.upper().strip() in name_upper:
            return val

    # 3. Filename client name is contained in config key
    for key, val in recips.items():
        if name_upper in key.upper().strip():
            return val

    return {"to": [], "cc": []}


# =============================================================================
# BULK RECIPIENT IMPORT FROM XLSX
# =============================================================================
#
# Lets users maintain the recipient list in Excel and bulk-import it into
# smartpayer_email_config.json instead of editing JSON by hand or re-keying
# everything in the GUI dialog.
#
# Expected XLSX schema (based on the user's List_RD_AR_Konfirmasi.xlsx):
#   - First sheet, first row may be blank (we hunt for the header row).
#   - Header columns (case-insensitive, any of these aliases work for NAME):
#       Account | City | TSM | NAME (or "Client Name", "Customer") |
#       Email To (or "To") | Email CC (or "CC")
#   - Each data row gives one recipient. Email cells can contain
#     semicolon-separated entries with optional "'Display Name' <addr>" or
#     "Display Name <addr>" framing — we extract just the bare addresses.
#
# Modes:
#   "merge"   — keep existing entries; new keys are added, existing keys are
#               OVERWRITTEN with the imported values (last-write-wins). This
#               is what the GUI button calls by default.
#   "replace" — wipe recipients dict, then write the imported entries.

# Email-address regex. Used to pull bare addresses out of strings like:
#   "'Display Name' <addr@dom.com>; 'Other' <other@dom.com>"
#   "addr@dom.com; bare2@dom.com"
#   "Display <a@b.c>, leftover@x.y"
# Tolerant of common separators ( ; , whitespace ) and of stray quotes.
_EMAIL_RE = re.compile(
    r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
)


def _extract_emails(cell_value) -> list:
    """
    Pull every email address out of a cell value. Returns a de-duplicated
    list preserving first-seen order. Returns [] for None / empty / non-str
    cells. Lowercase is NOT applied — many users want display-cased domains
    preserved (e.g. ssc.ClaimRD@godrejcp.com).
    """
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    seen = set()
    out = []
    for m in _EMAIL_RE.findall(text):
        m = m.strip().strip("'\"")
        key = m.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(m)
    return out


# Aliases for the four columns we actually use. Lookup is case-insensitive
# and strips whitespace, so "Email To" / "email to" / "EmailTo " all match.
_NAME_ALIASES = ("name", "client", "client name", "customer", "customer name", "payer", "payer name")
_TO_ALIASES   = ("email to", "to", "email_to", "emails to", "to emails", "recipient", "recipients")
_CC_ALIASES   = ("email cc", "cc", "email_cc", "emails cc", "cc emails", "cc recipients")


def _norm_header(s) -> str:
    return str(s or "").strip().lower()


def _find_header_row(ws) -> tuple:
    """
    Hunt the first 10 rows for the one that looks like a header — i.e. has
    a name-aliased cell AND an email-to-aliased cell. Returns
    (row_idx, col_map) where col_map is {logical_key: 1-based-col} for the
    keys "name", "to", "cc" (cc is optional and may be absent).

    Falls back to row 1 if no obvious header is detected.
    """
    max_scan = min(10, ws.max_row)
    for r in range(1, max_scan + 1):
        headers = [_norm_header(ws.cell(row=r, column=c).value)
                   for c in range(1, ws.max_column + 1)]
        col_map = {}
        for ci, h in enumerate(headers, start=1):
            if h in _NAME_ALIASES and "name" not in col_map:
                col_map["name"] = ci
            elif h in _TO_ALIASES and "to" not in col_map:
                col_map["to"] = ci
            elif h in _CC_ALIASES and "cc" not in col_map:
                col_map["cc"] = ci
        if "name" in col_map and "to" in col_map:
            return r, col_map
    # Last-resort: assume row 1 with default column layout
    #   A=Account, B=City, C=TSM, D=NAME, E=Email To, F=Email CC
    return 1, {"name": 4, "to": 5, "cc": 6}


def import_recipients_from_xlsx(xlsx_path: Path,
                                mode: str = "merge") -> dict:
    """
    Read an XLSX of recipients and merge / replace them into the config.
    Returns a summary dict:
        {"added": N, "updated": N, "skipped": N, "total": N, "mode": str}
    Raises FileNotFoundError if the XLSX is missing.

    The config is loaded, mutated, and saved in-place — callers don't need
    to wrap this in their own save_config().
    """
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                       check=True)
        import openpyxl

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    hdr_row, cmap = _find_header_row(ws)
    name_col = cmap["name"]
    to_col   = cmap["to"]
    cc_col   = cmap.get("cc")

    cfg = load_config() or {}
    existing = cfg.setdefault("recipients", {})

    if mode == "replace":
        existing.clear()

    # Build a case-insensitive index of existing keys so we can detect updates
    # without losing the user's original casing/spelling.
    existing_index = {k.upper().strip(): k for k in existing.keys()}

    added = updated = skipped = 0

    for r in range(hdr_row + 1, ws.max_row + 1):
        name_raw = ws.cell(row=r, column=name_col).value
        if name_raw is None or not str(name_raw).strip():
            skipped += 1
            continue
        # Drop any common Excel-prefix junk
        name = str(name_raw).strip().rstrip(".").strip()
        # Normalise the dictionary key to UPPERCASE for the existing
        # case-insensitive recipient-lookup logic to keep working.
        key = name.upper().strip()
        if not key:
            skipped += 1
            continue

        to_emails = _extract_emails(ws.cell(row=r, column=to_col).value)
        cc_emails = _extract_emails(ws.cell(row=r, column=cc_col).value) if cc_col else []

        if not to_emails:
            # A row with no usable TO email is useless — skip but log.
            skipped += 1
            continue

        entry = {"to": to_emails, "cc": cc_emails}

        if key in existing_index:
            existing[existing_index[key]] = entry
            updated += 1
        else:
            existing[key] = entry
            existing_index[key] = key
            added += 1

    save_config(cfg)

    summary = {
        "added":   added,
        "updated": updated,
        "skipped": skipped,
        "processed": added + updated,
        "mode":    mode,
    }
    return summary


# =============================================================================
# FILENAME PARSER
# =============================================================================

def parse_pdf_filename(pdf_path: Path) -> dict | None:
    """
    Parse a SmartPayer letter PDF filename and return its components.

    Handles these patterns:
      Smartpayer_April_2026_AGUS_SETIAWAN_letter.pdf
      SmartPayer April 2026 BAHARUDDIN_letter.pdf
      SmartPayer_April_2026_PT_PUTRA_NIAGA_SEJAHTERA_letter.pdf

    Returns: {month, year, client_name, next_month, next_year}
    or None if the filename doesn't match.
    """
    stem = pdf_path.stem   # e.g. "Smartpayer_April_2026_AGUS_SETIAWAN_letter"

    # Normalise separators: replace underscores with spaces
    stem_norm = stem.replace("_", " ")

    # Pattern: SmartPayer? {Month} {4-digit Year} {CLIENT NAME} letter
    pattern = re.compile(
        r'^[Ss]mart\s*[Pp]ayer\s+'
        r'([A-Za-z]+)\s+'          # group 1: month
        r'(\d{4})\s+'              # group 2: year
        r'(.+?)\s+letter$',        # group 3: client name
        re.IGNORECASE
    )
    m = pattern.match(stem_norm)
    if not m:
        return None

    month      = m.group(1).title()
    year       = m.group(2)
    client     = m.group(3).strip()
    nxt_month, nxt_year = _next_month(month, year)

    return {
        "month":       month,
        "year":        year,
        "client_name": client,
        "next_month":  nxt_month,
        "next_year":   nxt_year,
    }


# =============================================================================
# EMAIL COMPOSITION
# =============================================================================

def compose_email(cfg: dict, parsed: dict, pdf_path: Path) -> dict:
    """
    Build the full email dict: subject, body, to, cc, attachment.
    """
    tmpl    = cfg.get("email_template", {})
    cc_always = cfg.get("cc_always", [])
    recip   = get_recipient(cfg, parsed["client_name"])

    # Subject
    subject = tmpl.get(
        "subject",
        "Smart Payer {client_name} Periode {month} {year}"
    ).format(**parsed)

    # Body
    body = tmpl.get("body", "").format(**parsed)

    # Recipients
    to_list  = recip.get("to", [])
    cc_list  = list(recip.get("cc", [])) + [
        a for a in cc_always if a not in recip.get("cc", [])
    ]

    return {
        "subject":     subject,
        "body":        body,
        "to":          to_list,
        "cc":          cc_list,
        "attachment":  pdf_path,
    }


# =============================================================================
# SEND VIA SMTP
# =============================================================================

def send_smtp(cfg: dict, email: dict, account: dict = None) -> bool:
    smtp_cfg = (account or get_account(cfg)).get("smtp", {})
    host     = smtp_cfg.get("host", "")
    port     = int(smtp_cfg.get("port", 587))
    use_tls  = smtp_cfg.get("use_tls", True)
    username = smtp_cfg.get("username", "")
    password = smtp_cfg.get("password", "")
    sender_name  = smtp_cfg.get("sender_name", "ARRD Team")
    sender_email = smtp_cfg.get("sender_email", username)

    if not host or not username or not password:
        return False   # SMTP not configured — fall through to Outlook

    msg = MIMEMultipart()
    msg["From"]    = f"{sender_name} <{sender_email}>"
    msg["To"]      = ", ".join(email["to"])
    msg["Cc"]      = ", ".join(email["cc"])
    msg["Subject"] = email["subject"]
    msg.attach(MIMEText(email["body"], "plain", "utf-8"))

    # Attach PDF
    pdf_path = email["attachment"]
    with open(pdf_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{pdf_path.name}"')
    msg.attach(part)

    all_rcpt = email["to"] + email["cc"]
    try:
        if use_tls:
            with smtplib.SMTP(host, port, timeout=30) as srv:
                srv.ehlo()
                srv.starttls()
                srv.ehlo()
                srv.login(username, password)
                srv.sendmail(sender_email, all_rcpt, msg.as_bytes())
        else:
            with smtplib.SMTP_SSL(host, port, timeout=30) as srv:
                srv.login(username, password)
                srv.sendmail(sender_email, all_rcpt, msg.as_bytes())
        return True
    except Exception as e:
        _print(f"  [!] SMTP error: {e}")
        return False


# =============================================================================
# SEND VIA OUTLOOK COM (Windows only, no credentials needed)
# =============================================================================

def _account_smtp_address(acc) -> str:
    """
    Best-effort SMTP address for an Outlook Account COM object.

    For Exchange accounts, .SmtpAddress is sometimes blank (a known Outlook
    COM quirk) even though the account clearly has an SMTP address in the
    UI — in that case, resolve it via the Exchange user's primary SMTP
    address instead.
    """
    try:
        addr = str(acc.SmtpAddress or "").strip()
        if addr:
            return addr
    except Exception:
        pass
    try:
        exch_user = acc.CurrentUser.AddressEntry.GetExchangeUser()
        if exch_user is not None:
            return str(exch_user.PrimarySmtpAddress or "").strip()
    except Exception:
        pass
    return ""


def _find_outlook_account(outlook, smtp_address: str):
    """
    Look up an Outlook.Session.Accounts entry by SMTP address (case-insensitive).
    Returns the Account COM object, or None if no match / no address given.
    """
    if not smtp_address:
        return None
    target = smtp_address.strip().lower()
    try:
        for acc in outlook.Session.Accounts:
            try:
                if _account_smtp_address(acc).lower() == target:
                    return acc
            except Exception:
                continue
    except Exception:
        pass
    return None


def list_outlook_accounts():
    """
    Diagnostic: print every account Outlook automation can see, and how each
    one's SMTP address resolves. Run with --list-outlook-accounts on the
    Windows machine that has Outlook installed — this can't be tested from
    here since it requires live Outlook COM.
    """
    try:
        import win32com.client as win32
    except ImportError:
        _print("  [!] win32com not available. Run: pip install pywin32")
        return
    try:
        outlook = win32.Dispatch("Outlook.Application")
        accounts = list(outlook.Session.Accounts)
    except Exception as e:
        _print(f"  [!] Could not reach Outlook: {e}")
        return
    if not accounts:
        _print("  No accounts visible to Outlook automation.")
        return
    _print(f"  {len(accounts)} account(s) visible to Outlook automation:")
    for acc in accounts:
        try:
            name = acc.DisplayName
        except Exception:
            name = "(unknown)"
        resolved = _account_smtp_address(acc)
        try:
            raw = str(acc.SmtpAddress or "")
        except Exception:
            raw = "(error)"
        flag = "" if resolved else "  [!] could not resolve an SMTP address for this account"
        _print(f"    - {name}: SmtpAddress='{raw}'  resolved='{resolved}'{flag}")


def _create_mail_in_account_store(outlook, account_com_obj):
    """
    Create a new MailItem bound directly to the given account's own mail
    store, instead of Application.CreateItem(0) (which always binds to the
    profile's primary store first).

    This matters because MailItem.SendUsingAccount, set *after* creation,
    is known to silently fail to "stick" through .Send() on some Outlook
    builds when the item was created generically — Outlook still delivers
    from the store the item was originally created in, even though the
    property read-back looks correct. Creating the item via the target
    account's own default folder avoids that failure mode entirely, since
    the item is never associated with the wrong store in the first place.

    Returns the MailItem, or None if this account doesn't expose a usable
    DeliveryStore (falls back to the generic path in that case).
    """
    try:
        store = account_com_obj.DeliveryStore
        inbox = store.GetDefaultFolder(6)   # 6 = olFolderInbox
        return inbox.Items.Add("IPM.Note")
    except Exception:
        return None


def send_outlook(cfg: dict, email: dict, account: dict = None) -> bool:
    """
    Use the locally installed Outlook desktop app via win32com.
    Opens a new mail item pre-filled with all fields, then sends it.
    Outlook must already be signed in.

    IMPORTANT: with more than one Exchange account configured in Outlook,
    CreateItem(0) does NOT reliably honor the "Set as Default" account from
    Outlook's Account Settings UI — that setting only affects manually
    composed mail. Worse, MailItem.SendUsingAccount set on a generically
    created item can silently fail to take effect at Send() time on some
    Outlook builds, even though the property reads back correctly right
    after assignment. To reliably control the sending account, the mail
    item is created directly inside the target account's own mail store
    (see _create_mail_in_account_store) whenever that account is resolved,
    resolved from:
      1. account["outlook_fallback"]["send_as_account"]   (explicit override)
      2. account["smtp"]["sender_email"]                  (reuse SMTP identity)
    If neither is configured, or the address isn't found among the Outlook
    profile's accounts, this falls back to Outlook's own default behavior.
    """
    account = account or get_account(cfg)
    try:
        import win32com.client as win32
    except ImportError:
        try:
            subprocess.run(
                [sys.executable, "-m", "pip", "install", "pywin32", "-q"],
                check=True
            )
            import win32com.client as win32
        except Exception:
            _print("  [!] win32com not available. Install pywin32 or configure SMTP.")
            return False

    try:
        # Use early binding (gencache) rather than plain Dispatch(). Under
        # dynamic/late binding, assigning MailItem.SendUsingAccount can
        # silently no-op on some Windows/Outlook setups — no exception, the
        # property just doesn't stick, and the mail goes out from whatever
        # account Outlook considers default. Early binding resolves
        # Outlook's real type library and avoids that failure mode. Falls
        # back to plain Dispatch if gencache can't build (e.g. no write
        # access to the gencache dir).
        try:
            outlook = win32.gencache.EnsureDispatch("Outlook.Application")
        except Exception as e:
            _print(f"  [!] gencache.EnsureDispatch failed ({e}); falling back to "
                   f"Dispatch() — SendUsingAccount may not stick on this machine.")
            outlook = win32.Dispatch("Outlook.Application")

        send_as = (account.get("outlook_fallback", {}).get("send_as_account")
                   or account.get("smtp", {}).get("sender_email"))

        mail = None
        matched_account = None
        if send_as:
            matched_account = _find_outlook_account(outlook, send_as)
            if matched_account is not None:
                mail = _create_mail_in_account_store(outlook, matched_account)
                if mail is not None:
                    _print(f"    Created mail directly in account's store: "
                           f"{_account_smtp_address(matched_account)}")
                else:
                    _print("  [!] Could not create mail in the account's own store "
                           "(DeliveryStore unavailable) — falling back to "
                           "SendUsingAccount on a generically created item.")
            else:
                seen = [_account_smtp_address(a) or "(unresolved)" for a in outlook.Session.Accounts]
                _print(f"  [!] Outlook account '{send_as}' not found in this "
                       f"profile — falling back to Outlook's default account. "
                       f"(Outlook > File > Account Settings must list it.)")
                _print(f"      Accounts Outlook automation can see: {', '.join(seen) or '(none)'}")

        if mail is None:
            mail = outlook.CreateItem(0)   # 0 = olMailItem
            if matched_account is not None:
                # Belt-and-suspenders: still set it even though this path is
                # the less reliable one — it's a correct no-op if it doesn't
                # stick, and does work on some Outlook builds.
                mail.SendUsingAccount = matched_account
                _print(f"    Set SendUsingAccount to: "
                       f"{_account_smtp_address(matched_account)} "
                       f"(fallback path — verify the sent item's From address)")

        mail.Subject = email["subject"]
        mail.Body    = email["body"]
        mail.To      = "; ".join(email["to"])
        mail.CC      = "; ".join(email["cc"])

        # Attach PDF
        mail.Attachments.Add(str(email["attachment"].resolve()))

        mail.Send()
        return True
    except Exception as e:
        _print(f"  [!] Outlook COM error: {e}")
        return False


# =============================================================================
# MAIN SEND FUNCTION
# =============================================================================

def send_letter(pdf_path: Path, cfg: dict = None, log_fn=None, account: str = None) -> bool:
    """
    Parse the PDF filename, compose the email, and send it.
    Returns True on success.
    log_fn: optional callable(msg, tag) for GUI log panel.
    account: optional account name to send as, overriding cfg["active_account"]
             for this call only (used by --account / a GUI account picker).
    """
    def log(msg, tag=None):
        _print(f"  {msg}")
        if log_fn:
            log_fn(msg, tag)

    if cfg is None:
        cfg = load_config()
    if not cfg:
        log("[!] No email config loaded — skipping send.", "warning")
        return False

    acct_name = account or cfg.get("active_account")
    acct = get_account(cfg, acct_name)
    if not acct:
        log(f"[!] Account '{acct_name}' not found in config — skipping send.", "warning")
        log(f"    Known accounts: {', '.join(list_accounts(cfg)) or '(none)'}", "warning")
        return False
    log(f"    Sending as account: {acct_name}")

    # 1. Parse filename
    parsed = parse_pdf_filename(pdf_path)
    if not parsed:
        log(f"[!] Could not parse filename: {pdf_path.name}", "warning")
        log("    Expected: Smartpayer_Month_Year_CLIENT_NAME_letter.pdf", "warning")
        return False

    log(f"[>] Composing email for: {parsed['client_name']} ({parsed['month']} {parsed['year']})")

    # 2. Look up recipient
    recip = get_recipient(cfg, parsed["client_name"])
    if not recip.get("to"):
        log(f"[!] No recipient found for '{parsed['client_name']}'.", "warning")
        log("    Add them to smartpayer_email_config.json → recipients.", "warning")
        return False

    log(f"    TO:  {', '.join(recip['to'])}")
    cc = list(recip.get("cc", [])) + cfg.get("cc_always", [])
    if cc:
        log(f"    CC:  {', '.join(cc)}")

    # 3. Compose
    email = compose_email(cfg, parsed, pdf_path)

    # 4. Send — try SMTP first, fall back to Outlook
    smtp_cfg = acct.get("smtp", {})
    has_smtp = bool(smtp_cfg.get("host") and smtp_cfg.get("username")
                    and smtp_cfg.get("password"))

    if has_smtp:
        log("    Sending via SMTP...")
        if send_smtp(cfg, email, account=acct):
            log(f"[OK] Email sent to {', '.join(email['to'])}", "success")
            return True
        log("    SMTP failed — trying Outlook fallback...", "warning")

    if acct.get("outlook_fallback", {}).get("enabled", True):
        log("    Sending via Outlook...")
        if send_outlook(cfg, email, account=acct):
            log(f"[OK] Email sent via Outlook to {', '.join(email['to'])}", "success")
            return True

    log("[X] All send methods failed. Check SMTP config or Outlook installation.", "error")
    return False


# =============================================================================
# WATCH MODE
# =============================================================================

def watch_output_folder(folder: Path, log_fn=None, account: str = None):
    """
    Watch a folder for new *_letter.pdf files and auto-send each one.
    account: optional account name to send as (defaults to cfg["active_account"]).
    """
    def log(msg, tag=None):
        _print(f"  {msg}")
        if log_fn:
            log_fn(msg, tag)

    cfg = load_config()
    processed = set()

    log(f"[WATCH] Watching for new PDFs: {folder}")
    log("  Press Ctrl+C to stop.")

    try:
        while True:
            for pdf in Path(folder).glob("*_letter.pdf"):
                if pdf not in processed:
                    processed.add(pdf)
                    time.sleep(1)   # let file finish writing
                    log(f"\n[NEW] {pdf.name}", "info")
                    send_letter(pdf, cfg=cfg, log_fn=log_fn, account=account)
            time.sleep(3)
    except KeyboardInterrupt:
        log("\n[WATCH] Stopped.")


# =============================================================================
# CONFIG EDITOR (interactive CLI)
# =============================================================================

def _edit_account_cli(cfg: dict, name: str):
    """Prompt for one account's SMTP settings and write them into cfg['accounts'][name]."""
    def ask(prompt, current):
        val = input(f"  {prompt} [{current}]: ").strip()
        return val if val else current

    acct = cfg.setdefault("accounts", {}).setdefault(name, {})
    smtp = acct.setdefault("smtp", {})
    smtp["host"]         = ask("SMTP host (e.g. smtp.gmail.com)",    smtp.get("host","smtp.gmail.com"))
    smtp["port"]         = int(ask("SMTP port (587=TLS, 465=SSL)",   str(smtp.get("port",587))))
    smtp["use_tls"]      = ask("Use STARTTLS? (true/false)",         str(smtp.get("use_tls",True)).lower()) == "true"
    smtp["username"]     = ask("SMTP username / email",              smtp.get("username",""))
    smtp["password"]     = ask("SMTP password / app password",       smtp.get("password",""))
    smtp["sender_name"]  = ask("Sender display name",                smtp.get("sender_name","ARRD Team"))
    smtp["sender_email"] = ask("Sender email (blank = username)",    smtp.get("sender_email",""))
    if not smtp["sender_email"]:
        smtp["sender_email"] = smtp["username"]

    outlook = acct.setdefault("outlook_fallback", {})
    outlook["enabled"] = ask("Enable Outlook fallback for this account? (true/false)",
                              str(outlook.get("enabled", True)).lower()) == "true"
    outlook["send_as_account"] = ask(
        "Outlook 'send as' address (blank = reuse sender email above)",
        outlook.get("send_as_account", "")
    )


def _manage_accounts_cli(cfg: dict):
    """Add, edit, delete, or switch between named sender accounts."""
    accounts = cfg.setdefault("accounts", {})
    if not accounts:
        _print("\n  No accounts yet — let's create one.")
        name = input("  Account name (e.g. 'arrd', 'finance'): ").strip() or "default"
        _edit_account_cli(cfg, name)
        cfg["active_account"] = name
        return

    while True:
        _print("\n-- Accounts --")
        for n in accounts:
            marker = " (active)" if n == cfg.get("active_account") else ""
            smtp = accounts[n].get("smtp", {})
            _print(f"    {n}{marker}: {smtp.get('sender_email') or smtp.get('username') or '(not set)'}")
        _print("  [a]dd new  [e]dit existing  [s]witch active  [d]elete  [ENTER] done")
        choice = input("  Choice: ").strip().lower()

        if not choice:
            break
        elif choice == "a":
            name = input("  New account name: ").strip()
            if name:
                _edit_account_cli(cfg, name)
                if len(accounts) == 1 or not cfg.get("active_account"):
                    cfg["active_account"] = name
        elif choice == "e":
            name = input("  Account name to edit: ").strip()
            if name in accounts:
                _edit_account_cli(cfg, name)
            else:
                _print(f"  [!] No such account: {name}")
        elif choice == "s":
            name = input("  Account name to make active: ").strip()
            if set_active_account(cfg, name):
                _print(f"  [OK] Active account is now '{name}'")
            else:
                _print(f"  [!] No such account: {name}")
        elif choice == "d":
            name = input("  Account name to delete: ").strip()
            if name in accounts:
                del accounts[name]
                if cfg.get("active_account") == name:
                    cfg["active_account"] = next(iter(accounts), None)
                _print(f"  [OK] Deleted '{name}'")
            else:
                _print(f"  [!] No such account: {name}")


def edit_config_cli():
    """Interactive CLI to set up / edit the email config."""
    cfg = load_config() if CONFIG_FILE.exists() else {}

    _print("\n=== SmartPayer Email Config Setup ===\n")
    _print("Press ENTER to keep existing value.\n")

    def ask(prompt, current):
        val = input(f"  {prompt} [{current}]: ").strip()
        return val if val else current

    _manage_accounts_cli(cfg)

    _print("\n-- CC (always added to every email) --")
    cc_str = ask("CC addresses (comma-separated)", ", ".join(cfg.get("cc_always", [])))
    cfg["cc_always"] = [a.strip() for a in cc_str.split(",") if a.strip()]

    _print("\n-- Recipients --")
    recips = cfg.setdefault("recipients", {})
    _print("  Existing entries:")
    for name, val in recips.items():
        _print(f"    {name}: TO={val.get('to',[])}  CC={val.get('cc',[])}")
    while True:
        client = input("\n  Add/update client name (or ENTER to finish): ").strip().upper()
        if not client:
            break
        to_str = input(f"    TO emails for {client} (comma-separated): ").strip()
        cc_str = input(f"    CC emails for {client} (leave blank for none): ").strip()
        recips[client] = {
            "to": [a.strip() for a in to_str.split(",") if a.strip()],
            "cc": [a.strip() for a in cc_str.split(",") if a.strip()],
        }

    save_config(cfg)
    _print(f"\n[OK] Config saved to {CONFIG_FILE}")


# =============================================================================
# CLI ENTRY POINT
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="SmartPayer Email Sender")
    parser.add_argument("--pdf",         help="Send email for a specific PDF file")
    parser.add_argument("--watch",       help="Watch a folder for new letter PDFs and auto-send")
    parser.add_argument("--edit-config", action="store_true", help="Interactive config setup")
    parser.add_argument("--test-parse",  help="Test filename parsing without sending")
    parser.add_argument("--import-recipients", metavar="XLSX",
                        help="Bulk-import recipients from an XLSX file (columns: NAME, Email To, Email CC)")
    parser.add_argument("--import-mode", choices=("merge", "replace"), default="merge",
                        help="merge: add/overwrite per row (default). replace: wipe existing first.")
    parser.add_argument("--send-all",    metavar="DIR",
                        help="Send every *_letter.pdf in DIR (used by the GUI's test auto-mailer)")
    parser.add_argument("--account",     metavar="NAME",
                        help="Send using this account instead of the config's active_account")
    parser.add_argument("--list-accounts", action="store_true",
                        help="List configured sender accounts and exit")
    parser.add_argument("--set-active-account", metavar="NAME",
                        help="Set the default sender account and exit")
    parser.add_argument("--list-outlook-accounts", action="store_true",
                        help="Diagnostic: list accounts Outlook automation can see and how "
                             "their SMTP addresses resolve (Windows only, Outlook must be open)")
    args = parser.parse_args()

    if args.list_outlook_accounts:
        list_outlook_accounts()
    elif args.list_accounts:
        cfg = load_config()
        names = list_accounts(cfg)
        if not names:
            _print("No accounts configured. Run --edit-config to add one.")
        else:
            for n in names:
                marker = " (active)" if n == cfg.get("active_account") else ""
                smtp = get_account(cfg, n).get("smtp", {})
                _print(f"  {n}{marker}: {smtp.get('sender_email') or smtp.get('username') or '(not set)'}")
    elif args.set_active_account:
        cfg = load_config()
        if set_active_account(cfg, args.set_active_account):
            save_config(cfg)
            _print(f"[OK] Active account set to '{args.set_active_account}'")
        else:
            _print(f"[X] No such account: {args.set_active_account}. "
                   f"Known: {', '.join(list_accounts(cfg)) or '(none)'}")
            sys.exit(1)
    elif args.edit_config:
        edit_config_cli()
    elif args.import_recipients:
        try:
            summary = import_recipients_from_xlsx(Path(args.import_recipients),
                                                  mode=args.import_mode)
        except Exception as e:
            _print(f"[X] Import failed: {e}")
            sys.exit(1)
        _print(f"[OK] Imported recipients (mode={summary['mode']}): "
               f"added={summary['added']}, updated={summary['updated']}, "
               f"skipped={summary['skipped']}, processed={summary['processed']}")
    elif args.send_all:
        folder = Path(args.send_all)
        if not folder.is_dir():
            _print(f"[X] Not a folder: {folder}")
            sys.exit(1)
        cfg = load_config()
        pdfs = sorted(folder.glob("*_letter.pdf"))
        if not pdfs:
            _print(f"[!] No *_letter.pdf files found in {folder}")
            sys.exit(0)
        _print(f"[*] Sending {len(pdfs)} PDF(s) from {folder}...")
        ok = fail = 0
        for pdf in pdfs:
            _print(f"\n--- {pdf.name} ---")
            try:
                if send_letter(pdf, cfg=cfg, account=args.account):
                    ok += 1
                else:
                    fail += 1
            except Exception as e:
                _print(f"[X] Error sending {pdf.name}: {e}")
                fail += 1
        _print(f"\n[DONE] {ok} sent, {fail} failed.")
        sys.exit(0 if fail == 0 else 1)
    elif args.test_parse:
        p = parse_pdf_filename(Path(args.test_parse))
        _print(json.dumps(p, indent=2) if p else "Could not parse filename.")
    elif args.pdf:
        success = send_letter(Path(args.pdf), account=args.account)
        sys.exit(0 if success else 1)
    elif args.watch:
        watch_output_folder(Path(args.watch), account=args.account)
    else:
        _print("Usage: python smartpayer_emailer.py --pdf <file.pdf> [--account NAME]")
        _print("       python smartpayer_emailer.py --watch <output_folder> [--account NAME]")
        _print("       python smartpayer_emailer.py --send-all <output_folder> [--account NAME]")
        _print("       python smartpayer_emailer.py --import-recipients <xlsx>")
        _print("       python smartpayer_emailer.py --edit-config")
        _print("       python smartpayer_emailer.py --list-accounts")
        _print("       python smartpayer_emailer.py --set-active-account <NAME>")
        _print("       python smartpayer_emailer.py --list-outlook-accounts")
        _print("       python smartpayer_emailer.py --test-parse <filename.pdf>")


if __name__ == "__main__":
    main()
