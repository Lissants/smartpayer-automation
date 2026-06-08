"""
smartpayer_automation.py  -  v8
Core logic. Called by smartpayer_gui.py via run_automation().
Changes from v7:
  - Permission denied errors on Smartpayer.xlsx are caught cleanly.
    The script waits and retries up to 10 times (30 seconds total),
    prompting the user via progress_cb to close the file in Excel.
"""

import os, sys, re, subprocess, warnings, glob, shutil, zipfile, time
from datetime import datetime, date, timedelta
from collections import defaultdict

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning,
                        message=".*Data Validation extension.*")

SMARTPAYER_FILE = "Smartpayer.xlsx"
OUTPUT_FOLDER   = "OUTPUT"
BACKUP_FOLDER   = "BACKUP"
INPUT_FOLDER    = "INPUT"
MASTER_SHEET    = "Master"

BILL_TO_GROUP_PAYERS = {
    "CV. INDRA JAYA",
    "CV. SEJATI MANDIRI - BAU BAU",
}

# Master column positions (0-based)
M_PAYER      = 3
M_BILL_TO    = 4
M_PAYER_NAME = 5
M_BILL_TO_NM = 6
M_DOC_NO     = 8
M_DOC_DATE   = 10
M_NET_DUE    = 11
M_TOT_INV    = 13

# Sheet1 output columns (1-based)
C_PAYER      = 1
C_BILL_TO    = 2
C_PAYER_NAME = 3
C_DOC_NO     = 4
C_DOC_DATE   = 5
C_NET_DUE    = 6
C_TOT_INV    = 7
C_DPP        = 8
C_DATE1      = 9

RATE_COL    = 15
HDR_ROW     = 1
DATA_START  = 2

C_BLACK     = "FF000000"
C_WHITE     = "FFFFFFFF"
C_ORANGE    = "FFF4B183"
C_GREEN_L80 = "FFA9D18E"
C_DARK_BLUE = "FF203864"
C_YELLOW    = "FFFFFF00"


def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_fill():   return mk_fill(C_BLACK)
def hdr_font():   return Font(color=C_WHITE, bold=True, name="Calibri")
def hdr_align():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left_align(): return Alignment(horizontal="left", vertical="center")
def money():      return "#,##0"

def all_borders():
    s = Side(style="thin", color=C_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_summary_style(cell, row_type):
    if row_type == "pph":
        cell.fill = mk_fill(C_ORANGE)
        cell.font = Font(color=C_BLACK, bold=True, name="Calibri")
    elif row_type == "net":
        cell.fill = mk_fill(C_GREEN_L80)
        cell.font = Font(color=C_BLACK, bold=True, name="Calibri")
    elif row_type == "nom":
        cell.fill = mk_fill(C_DARK_BLUE)
        cell.font = Font(color=C_WHITE, bold=True, name="Calibri")

def col(idx):
    return get_column_letter(idx)


# =============================================================================
# SAFE FILE OPERATIONS  (retry on Permission Denied)
# =============================================================================

def _safe_load(filepath, progress_cb, data_only=False, max_retries=10, wait=3):
    """
    Load a workbook, retrying if the file is locked by Excel.
    Raises PermissionError after max_retries attempts.
    """
    for attempt in range(1, max_retries + 1):
        try:
            return load_workbook(filepath, data_only=data_only)
        except PermissionError:
            if attempt == max_retries:
                raise PermissionError(
                    f"Cannot open '{filepath}' - it is still open in Excel.\n"
                    f"Please close the file and try again."
                )
            progress_cb(
                f"WARNING: '{os.path.basename(filepath)}' is open in Excel. "
                f"Please close it. Retrying in {wait}s... "
                f"(attempt {attempt}/{max_retries})"
            )
            time.sleep(wait)


def _safe_save(wb, filepath, progress_cb, max_retries=10, wait=3):
    """
    Save a workbook, retrying if the file is locked by Excel.
    Raises PermissionError after max_retries attempts.
    """
    for attempt in range(1, max_retries + 1):
        try:
            wb.save(filepath)
            return
        except PermissionError:
            if attempt == max_retries:
                raise PermissionError(
                    f"Cannot save '{filepath}' - it is still open in Excel.\n"
                    f"Please close the file and try again."
                )
            progress_cb(
                f"WARNING: '{os.path.basename(filepath)}' is open in Excel. "
                f"Please close it. Retrying in {wait}s... "
                f"(attempt {attempt}/{max_retries})"
            )
            time.sleep(wait)


# =============================================================================
# MONTH NAMING HELPER
# =============================================================================

def get_month_year_label(filtered_rows):
    """
    Determine <Month> <Year> for file naming.
    Rule: one month BEFORE the most common Net Due Date month.
    """
    from collections import Counter
    month_counts = Counter()
    for rec in filtered_rows:
        nd = rec.get("net_due")
        if nd is None:
            continue
        if isinstance(nd, (datetime, date)):
            month_counts[(nd.year, nd.month)] += 1

    if not month_counts:
        today = date.today()
        m = today.month - 1 or 12
        y = today.year if today.month > 1 else today.year - 1
        return datetime(y, m, 1).strftime("%B %Y")

    (y, m), _ = month_counts.most_common(1)[0]
    if m == 1:
        m, y = 12, y - 1
    else:
        m -= 1
    return datetime(y, m, 1).strftime("%B %Y")


# =============================================================================
# OUTPUT FOLDER ARCHIVER
# =============================================================================

def archive_output_folder(month_year, progress_cb):
    """
    Compress all existing .xlsx files in the OUTPUT folder into a single
    .zip file named "Output Smartpayer <Month> <Year>.zip", move it to
    BACKUP, then delete the individual files from OUTPUT.
    Does nothing if OUTPUT is empty or does not exist.
    """
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    existing = [
        f for f in glob.glob(os.path.join(OUTPUT_FOLDER, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]

    if not existing:
        progress_cb("OUTPUT folder is empty - nothing to archive.")
        return None

    zip_name = f"Output Smartpayer {month_year}.zip"
    zip_tmp  = os.path.join(OUTPUT_FOLDER, zip_name)
    zip_dest = os.path.join(BACKUP_FOLDER, zip_name)

    progress_cb(f"Archiving {len(existing)} existing output file(s) into {zip_name}...")

    with zipfile.ZipFile(zip_tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for filepath in sorted(existing):
            zf.write(filepath, arcname=os.path.basename(filepath))

    for filepath in existing:
        os.remove(filepath)

    if os.path.exists(zip_dest):
        os.remove(zip_dest)
    shutil.move(zip_tmp, zip_dest)

    progress_cb(f"Archive saved to BACKUP: {zip_name}")
    return zip_dest


# =============================================================================
# INPUT FILE HANDLING
# =============================================================================

def find_input_file():
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    files = glob.glob(os.path.join(INPUT_FOLDER, "*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    return files[0]


def copy_master_from_input(input_path, progress_cb):
    progress_cb("Reading input file...")
    wb_input = _safe_load(input_path, progress_cb)
    ws_input = wb_input.active

    progress_cb("Opening Smartpayer.xlsx to update Master sheet...")
    wb_main = _safe_load(SMARTPAYER_FILE, progress_cb)

    if MASTER_SHEET not in wb_main.sheetnames:
        wb_main.create_sheet(MASTER_SHEET)

    ws_master = wb_main[MASTER_SHEET]

    for row in ws_master.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    progress_cb("Copying data to Master sheet...")
    for r_idx, row in enumerate(ws_input.iter_rows(values_only=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws_master.cell(row=r_idx, column=c_idx).value = val

    _safe_save(wb_main, SMARTPAYER_FILE, progress_cb)
    wb_input.close()
    progress_cb("Master sheet updated.")
    return wb_main


def archive_input_file(input_path, month_year, move=True):
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    dest_name = f"Backup Master Data {month_year}.xlsx"
    dest_path = os.path.join(BACKUP_FOLDER, dest_name)
    if os.path.exists(dest_path):
        os.remove(dest_path)
    if move:
        shutil.move(input_path, dest_path)
    else:
        shutil.copy2(input_path, dest_path)
    return dest_path


# =============================================================================
# MASTER READER
# =============================================================================

def read_master(wb):
    ws = wb[MASTER_SHEET]
    all_rows = []
    filtered = []
    skipped  = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[M_PAYER] is None:
            continue
        tot = row[M_TOT_INV]
        rec = {
            "payer_id"    : row[M_PAYER],
            "bill_to"     : row[M_BILL_TO],
            "bill_to_name": row[M_BILL_TO_NM] if len(row) > M_BILL_TO_NM else None,
            "payer_name"  : str(row[M_PAYER_NAME]).strip() if row[M_PAYER_NAME] else "",
            "doc_no"      : row[M_DOC_NO],
            "doc_date"    : row[M_DOC_DATE],
            "net_due"     : row[M_NET_DUE],
            "tot_inv"     : tot if tot else 0,
        }
        all_rows.append(rec)
        if tot is None or tot < 0:
            skipped += 1
        else:
            filtered.append(rec)
    return all_rows, filtered, skipped


# =============================================================================
# DATE UTILITIES
# =============================================================================

def to_dt(v):
    if v is None:               return None
    if isinstance(v, datetime): return v
    if isinstance(v, date):     return datetime(v.year, v.month, v.day)
    return None


def net_due_window_from_payment_dates(dates):
    """
    Use the first selected payment date to determine the eligible Net Due Date
    month. Example: payment date 27-Apr-2026 -> 01-May-2026 through
    31-May-2026.
    """
    first_date = next((d for d in dates if d is not None), None)
    if first_date is None:
        return None, None

    if first_date.month == 12:
        start = datetime(first_date.year + 1, 1, 1)
        next_month = datetime(first_date.year + 1, 2, 1)
    else:
        start = datetime(first_date.year, first_date.month + 1, 1)
        if start.month == 12:
            next_month = datetime(start.year + 1, 1, 1)
        else:
            next_month = datetime(start.year, start.month + 1, 1)
    end = next_month - timedelta(microseconds=1)
    return start, end


def filter_by_net_due_window(rows, start, end):
    if start is None or end is None:
        return rows, 0
    kept = []
    skipped = 0
    for rec in rows:
        nd = to_dt(rec.get("net_due"))
        if nd and start <= nd <= end:
            kept.append(rec)
        else:
            skipped += 1
    return kept, skipped


# =============================================================================
# GROUPING HELPERS
# =============================================================================

def is_bill_to_grouped(payer_name):
    return payer_name.strip().upper() in {p.upper() for p in BILL_TO_GROUP_PAYERS}


def group_key(rec):
    pn = rec["payer_name"].strip()
    if is_bill_to_grouped(pn):
        return f"{pn}||{rec['bill_to']}"
    return pn


def group_label(key_str):
    if "||" in key_str:
        pn, bt = key_str.split("||", 1)
        return f"{pn} {bt} Total"
    return f"{key_str} Total"


def sort_key(rec):
    pn = rec["payer_name"].strip()
    if is_bill_to_grouped(pn):
        return (pn.upper(), str(rec["bill_to"] or "").upper(), str(rec["doc_no"] or ""))
    return (pn.upper(), "", str(rec["doc_no"] or ""))


# =============================================================================
# SHEET1 BUILDER
# =============================================================================

def build_sheet1(wb, filtered_rows, dates, daily_rate):
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]
    ws = wb.create_sheet("Sheet1", 0)

    use_d4   = dates[3] is not None
    n_dcols  = 4 if use_d4 else 3
    last_col = C_DATE1 + n_dcols - 1

    ws.cell(row=1, column=RATE_COL).value         = daily_rate
    ws.cell(row=1, column=RATE_COL).number_format = "0.00000%"
    ws.cell(row=2, column=RATE_COL).value         = "=O1*365"

    fixed_hdrs = [
        "Payer", "Bill To Party", "Payer Name", "Document No",
        "Doc. Date", "Net Due Date", "Tot.Inv.Ostd.", "DPP",
    ]
    for ci, h in enumerate(fixed_hdrs, 1):
        cell = ws.cell(row=HDR_ROW, column=ci)
        cell.value     = h
        cell.fill      = hdr_fill()
        cell.font      = hdr_font()
        cell.alignment = hdr_align()

    used_date_hdrs = []
    for i, d in enumerate(dates[:n_dcols]):
        if d:
            ci   = C_DATE1 + i
            cell = ws.cell(row=HDR_ROW, column=ci)
            cell.value         = datetime(d.year, d.month, d.day)
            cell.number_format = "DD-MMM-YYYY"
            cell.fill          = hdr_fill()
            cell.font          = hdr_font()
            cell.alignment     = hdr_align()
            used_date_hdrs.append(d)

    filtered_rows.sort(key=sort_key)

    current_row = DATA_START
    prev_key    = None
    group_first = {}
    group_last  = {}

    for rec in filtered_rows:
        gk = group_key(rec)

        if prev_key is not None and gk != prev_key:
            _write_subtotal_row(ws, current_row,
                                group_first[prev_key], group_last[prev_key],
                                prev_key, last_col, n_dcols)
            current_row += 1

        group_first.setdefault(gk, current_row)
        _write_data_row(ws, current_row, rec, n_dcols)
        group_last[gk] = current_row
        prev_key       = gk
        current_row   += 1

    if prev_key is not None:
        _write_subtotal_row(ws, current_row,
                            group_first[prev_key], group_last[prev_key],
                            prev_key, last_col, n_dcols)
        current_row += 1

    last_row = current_row - 1
    ws.auto_filter.ref = f"A{HDR_ROW}:{col(last_col)}{last_row}"
    ws.freeze_panes    = ws.cell(row=DATA_START, column=1)

    for ci, w in [(1,12),(2,14),(3,30),(4,14),(5,13),(6,13),(7,16),(8,16)]:
        ws.column_dimensions[col(ci)].width = w
    for i in range(n_dcols):
        ws.column_dimensions[col(C_DATE1+i)].width = 18

    return last_row, used_date_hdrs


def _write_data_row(ws, row, rec, n_dcols):
    ws.cell(row=row, column=C_PAYER     ).value = rec["payer_id"]
    ws.cell(row=row, column=C_BILL_TO   ).value = rec["bill_to"]
    ws.cell(row=row, column=C_PAYER_NAME).value = rec["payer_name"]
    ws.cell(row=row, column=C_DOC_NO    ).value = rec["doc_no"]

    for ci, key, fmt in [(C_DOC_DATE, "doc_date", "DD/MM/YYYY"),
                          (C_NET_DUE,  "net_due",  "DD/MM/YYYY")]:
        v = to_dt(rec[key])
        ws.cell(row=row, column=ci).value         = v
        ws.cell(row=row, column=ci).number_format = fmt

    ws.cell(row=row, column=C_TOT_INV).value         = rec["tot_inv"]
    ws.cell(row=row, column=C_TOT_INV).number_format = money()

    ws.cell(row=row, column=C_DPP).value         = f"={col(C_TOT_INV)}{row}/1.11"
    ws.cell(row=row, column=C_DPP).number_format = money()

    for i in range(n_dcols):
        out_ci  = C_DATE1 + i
        out_cl  = col(out_ci)
        formula = f"=({col(C_NET_DUE)}{row}-{out_cl}$1)*$O$1*{col(C_DPP)}{row}"
        ws.cell(row=row, column=out_ci).value         = formula
        ws.cell(row=row, column=out_ci).number_format = money()


def _write_subtotal_row(ws, row, first, last, gk, last_col, n_dcols):
    label = group_label(gk)
    for ci in range(1, last_col + 1):
        ws.cell(row=row, column=ci).fill = hdr_fill()
        ws.cell(row=row, column=ci).font = hdr_font()
    cell = ws.cell(row=row, column=C_PAYER_NAME)
    cell.value     = label
    cell.fill      = hdr_fill()
    cell.font      = hdr_font()
    cell.alignment = left_align()
    for ci in range(C_TOT_INV, last_col + 1):
        cl = col(ci)
        ws.cell(row=row, column=ci).value         = f"=SUM({cl}{first}:{cl}{last})"
        ws.cell(row=row, column=ci).number_format = money()


# =============================================================================
# SUMMARY TABLE
# =============================================================================

def build_summary_table(wb, last_data_row, filtered_rows, n_dcols):
    ws = wb["Sheet1"]
    pph_row = last_data_row + 5
    net_row = pph_row + 1
    nom_row = pph_row + 2

    HELPER_COL = 14
    for r in range(DATA_START, last_data_row + 1):
        pname_cell = ws.cell(row=r, column=C_PAYER_NAME).value
        bt_cell    = ws.cell(row=r, column=C_BILL_TO).value
        if pname_cell and "Total" not in str(pname_cell):
            if is_bill_to_grouped(str(pname_cell)):
                ws.cell(row=r, column=HELPER_COL).value = f"{pname_cell} {bt_cell}"
            else:
                ws.cell(row=r, column=HELPER_COL).value = pname_cell
        ws.cell(row=r, column=HELPER_COL).font = Font(color="FFFFFFFF")

    row_types = [("pph", "PPH 15%"),
                 ("net", "Net Disc"),
                 ("nom", "Nominal yang di transfer")]

    for i, (rtype, lbl) in enumerate(row_types):
        cell = ws.cell(row=pph_row + i, column=C_TOT_INV)
        cell.value     = lbl
        cell.alignment = left_align()
        apply_summary_style(cell, rtype)

    for i, (rtype, _) in enumerate(row_types):
        cell = ws.cell(row=pph_row + i, column=C_DPP)
        apply_summary_style(cell, rtype)

    payer_names = sorted(set(
        f"{r['payer_name']} {r['bill_to']}"
        if is_bill_to_grouped(r["payer_name"])
        else r["payer_name"]
        for r in filtered_rows if r["payer_name"]
    ))

    helper_name = "_PayerList"
    if helper_name in wb.sheetnames:
        del wb[helper_name]
    helper = wb.create_sheet(helper_name)
    helper.sheet_state = "hidden"
    for i, pn in enumerate(payer_names, 1):
        helper.cell(row=i, column=1).value = pn

    dv = DataValidation(
        type="list",
        formula1=f"={helper_name}!$A$1:$A${len(payer_names)}",
        allow_blank=False,
        showErrorMessage=False,
    )
    dv.sqref = f"C{pph_row}"
    ws.add_data_validation(dv)

    dropdown_cell = ws.cell(row=pph_row, column=C_PAYER_NAME)
    dropdown_cell.value     = payer_names[0]
    dropdown_cell.fill      = mk_fill(C_YELLOW)
    dropdown_cell.font      = Font(color=C_BLACK, bold=True, name="Calibri")
    dropdown_cell.border    = all_borders()
    dropdown_cell.alignment = left_align()

    choice       = f"$C${pph_row}"
    helper_range = f"$N${DATA_START}:$N${last_data_row}"
    totinv_range = f"$G${DATA_START}:$G${last_data_row}"

    for i in range(n_dcols):
        out_ci     = C_DATE1 + i
        out_cl     = col(out_ci)
        date_range = f"${out_cl}${DATA_START}:${out_cl}${last_data_row}"
        for row_off, (rtype, formula) in enumerate([
            ("pph", f"=SUMIF({helper_range},{choice},{date_range})*15%"),
            ("net", f"=SUMIF({helper_range},{choice},{date_range})-{out_cl}{pph_row}"),
            ("nom", f"=SUMIF({helper_range},{choice},{totinv_range})-{out_cl}{pph_row+1}"),
        ]):
            cell = ws.cell(row=pph_row + row_off, column=out_ci)
            cell.value         = formula
            cell.number_format = money()
            apply_summary_style(cell, rtype)

    return pph_row


# =============================================================================
# RECALCULATION
# =============================================================================

def recalculate(filepath):
    abs_path = os.path.abspath(filepath)
    if sys.platform == "win32":
        try:
            import win32com.client as win32
            xl = win32.gencache.EnsureDispatch("Excel.Application")
            xl.Visible = xl.DisplayAlerts = False
            wbc = xl.Workbooks.Open(abs_path)
            xl.Application.CalculateFull()
            wbc.Save()
            wbc.Close(False)
            xl.Quit()
            return True
        except Exception:
            return False
    return False


# =============================================================================
# OUTPUT FILES
# =============================================================================

def _output_filename(month_year, payer_name, bill_to=None):
    parts = ["Smartpayer", month_year, payer_name]
    if bill_to:
        parts.append(str(bill_to))
    name = " ".join(parts)
    return re.sub(r'[\\/:*?"<>|]', '', name).strip() + ".xlsx"


def _write_output_rows(out_ws, rows, payer_name, n_dcols, is_exception):
    """
    Write all data rows and subtotal row(s) into an output worksheet.
    Normal payers:    all rows then one subtotal at the end.
    Exception payers: rows grouped by Bill To Party, one subtotal per group.
    """
    last_col = C_DATE1 + n_dcols - 1
    out_row  = DATA_START

    if is_exception:
        rows.sort(key=lambda v: (str(v[C_BILL_TO-1] or "").upper(),
                                  str(v[C_DOC_NO-1] or "")))
        prev_bt   = None
        grp_first = {}
        grp_last  = {}

        for vals in rows:
            bt = str(vals[C_BILL_TO - 1] or "")
            if prev_bt is not None and bt != prev_bt:
                _out_subtotal_row(out_ws, out_row,
                                  grp_first[prev_bt], grp_last[prev_bt],
                                  payer_name, prev_bt, last_col, n_dcols)
                out_row += 1
            grp_first.setdefault(bt, out_row)
            _out_data_row(out_ws, out_row, vals, n_dcols)
            grp_last[bt] = out_row
            prev_bt      = bt
            out_row     += 1

        if prev_bt:
            _out_subtotal_row(out_ws, out_row,
                              grp_first[prev_bt], grp_last[prev_bt],
                              payer_name, prev_bt, last_col, n_dcols)
            out_row += 1

    else:
        rows.sort(key=lambda v: str(v[C_DOC_NO-1] or ""))
        data_first = out_row

        for vals in rows:
            _out_data_row(out_ws, out_row, vals, n_dcols)
            out_row += 1

        data_last = out_row - 1
        _out_subtotal_row(out_ws, out_row,
                          data_first, data_last,
                          payer_name, None, last_col, n_dcols)
        out_row += 1

    return out_row - 1


def _out_data_row(out_ws, row, vals, n_dcols):
    for ci in range(1, C_DATE1 + n_dcols):
        v    = vals[ci - 1]
        cell = out_ws.cell(row=row, column=ci)
        cell.value = v
        if ci in (C_DOC_DATE, C_NET_DUE):
            cell.number_format = "DD/MM/YYYY"
        elif ci >= C_TOT_INV:
            cell.number_format = money()


def _out_subtotal_row(ws, row, first, last, payer_name, bill_to,
                      last_col, n_dcols):
    label = (f"{payer_name} {bill_to} Total"
             if bill_to else f"{payer_name} Total")
    for ci in range(1, last_col + 1):
        ws.cell(row=row, column=ci).fill = hdr_fill()
        ws.cell(row=row, column=ci).font = hdr_font()
    ws.cell(row=row, column=C_PAYER_NAME).value     = label
    ws.cell(row=row, column=C_PAYER_NAME).alignment = left_align()
    for ci in range(C_TOT_INV, last_col + 1):
        cl = col(ci)
        ws.cell(row=row, column=ci).value         = f"=SUM({cl}{first}:{cl}{last})"
        ws.cell(row=row, column=ci).number_format = money()


def create_output_files(source_path, filtered_rows, dates, last_data_row,
                        n_dcols, daily_rate, month_year, progress_cb):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    wb_src = _safe_load(source_path, progress_cb, data_only=True)
    ws_src = wb_src["Sheet1"]

    all_src = []
    for r in range(DATA_START, last_data_row + 1):
        vals = [ws_src.cell(row=r, column=c).value
                for c in range(1, C_DATE1 + n_dcols)]
        all_src.append(vals)

    key_date_sums = defaultdict(lambda: defaultdict(float))
    key_inv_sums  = defaultdict(float)

    for vals in all_src:
        pname = vals[C_PAYER_NAME - 1]
        if not pname or "Total" in str(pname):
            continue
        pname = str(pname).strip()
        bt    = vals[C_BILL_TO - 1]
        agg_key = (f"{pname}||{bt}"
                   if is_bill_to_grouped(pname)
                   else pname)
        tot = vals[C_TOT_INV - 1]
        if isinstance(tot, (int, float)):
            key_inv_sums[agg_key] += tot
        for di in range(n_dcols):
            v = vals[C_DATE1 - 1 + di]
            if isinstance(v, (int, float)):
                key_date_sums[agg_key][di] += v

    tasks = {}
    for vals in all_src:
        pname = vals[C_PAYER_NAME - 1]
        if not pname or "Total" in str(pname):
            continue
        pname = str(pname).strip()
        bt    = vals[C_BILL_TO - 1]
        if is_bill_to_grouped(pname):
            agg_key = f"{pname}||{bt}"
            tasks.setdefault(agg_key, {"payer_name": pname, "bill_to": bt, "rows": []})
        else:
            agg_key = pname
            tasks.setdefault(agg_key, {"payer_name": pname, "bill_to": None, "rows": []})
        tasks[agg_key]["rows"].append(vals)

    created = []
    total   = len(tasks)

    for idx, (agg_key, task) in enumerate(sorted(tasks.items()), 1):
        payer_name   = task["payer_name"]
        bill_to      = task["bill_to"]
        rows         = task["rows"]
        is_exception = is_bill_to_grouped(payer_name)

        progress_cb(f"Creating output file {idx}/{total}: {payer_name}...")

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sheet1"

        out_ws.cell(row=1, column=RATE_COL).value         = daily_rate
        out_ws.cell(row=1, column=RATE_COL).number_format = "0.00000%"
        out_ws.cell(row=2, column=RATE_COL).value         = "Simulasi Imbalan"

        fixed_hdrs = ["Payer", "Bill To Party", "Payer Name", "Document No",
                       "Doc. Date", "Net Due Date", "Tot.Inv.Ostd.", "DPP"]
        for ci, h in enumerate(fixed_hdrs, 1):
            cell = out_ws.cell(row=HDR_ROW, column=ci)
            cell.value, cell.fill, cell.font, cell.alignment = (
                h, hdr_fill(), hdr_font(), hdr_align())
        for i, d in enumerate(dates[:n_dcols]):
            if d:
                ci = C_DATE1 + i
                cell = out_ws.cell(row=HDR_ROW, column=ci)
                cell.value         = datetime(d.year, d.month, d.day)
                cell.number_format = "DD-MMM-YYYY"
                cell.fill, cell.font, cell.alignment = (
                    hdr_fill(), hdr_font(), hdr_align())

        last_out = _write_output_rows(
            out_ws, rows, payer_name, n_dcols, is_exception)

        sum_start = last_out + 5
        row_types = [("pph", "PPH 15%"), ("net", "Net Disc"),
                     ("nom", "Nominal yang di transfer")]

        for i, (rtype, lbl) in enumerate(row_types):
            cell = out_ws.cell(row=sum_start + i, column=C_TOT_INV)
            cell.value     = lbl
            cell.alignment = left_align()
            apply_summary_style(cell, rtype)
            apply_summary_style(
                out_ws.cell(row=sum_start + i, column=C_DPP), rtype)

        date_sums = key_date_sums[agg_key]
        inv_sum   = key_inv_sums[agg_key]

        for di in range(n_dcols):
            date_sum = date_sums.get(di, 0.0)
            pph_val  = date_sum * 0.15
            net_val  = date_sum - pph_val
            nom_val  = inv_sum  - net_val
            out_ci   = C_DATE1 + di
            for ri, (rtype, v) in enumerate(
                    zip(["pph", "net", "nom"], [pph_val, net_val, nom_val])):
                cell = out_ws.cell(row=sum_start + ri, column=out_ci)
                cell.value         = round(v)
                cell.number_format = money()
                apply_summary_style(cell, rtype)

        for ci, w in [(1,12),(2,14),(3,30),(4,14),(5,13),(6,13),(7,16),(8,16)]:
            out_ws.column_dimensions[col(ci)].width = w
        for i in range(n_dcols):
            out_ws.column_dimensions[col(C_DATE1+i)].width = 18

        filename = _output_filename(month_year, payer_name, bill_to)
        out_path = os.path.join(OUTPUT_FOLDER, filename)
        out_wb.save(out_path)
        created.append(out_path)

    return created


# =============================================================================
# MAIN ENTRY POINT (called by GUI)
# =============================================================================

def run_automation(params, progress_cb):
    """
    params = {
        "daily_rate": float,
        "dates": [date, date, date, date|None],
        "input_path": optional explicit input workbook path
    }
    progress_cb(message: str)
    Returns list of created output file paths.
    Raises Exception on any error.
    """
    daily_rate = params["daily_rate"]
    dates      = params["dates"]
    use_d4     = dates[3] is not None
    n_dcols    = 4 if use_d4 else 3

    # Step 1: Find input file
    explicit_input = params.get("input_path")
    if explicit_input:
        progress_cb("Using selected input workbook...")
        input_path = explicit_input
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Selected input workbook not found:\n{input_path}")
    else:
        progress_cb("Looking for input file in INPUT folder...")
        input_path = find_input_file()
    if not input_path:
        raise FileNotFoundError(
            f"No .xlsx file found in the '{INPUT_FOLDER}' folder.\n"
            f"Please place your master data file there and try again."
        )
    progress_cb(f"Found input file: {os.path.basename(input_path)}")

    # Step 2: Copy master data
    wb = copy_master_from_input(input_path, progress_cb)

    # Step 3: Read master
    progress_cb("Reading and filtering Master data...")
    all_rows, filtered_rows, skipped = read_master(wb)
    due_start, due_end = net_due_window_from_payment_dates(dates)
    if due_start and due_end:
        filtered_rows, skipped_due = filter_by_net_due_window(
            filtered_rows, due_start, due_end
        )
        progress_cb(
            "Net Due Date filter: "
            f"{due_start.strftime('%d-%b-%Y')} to {due_end.strftime('%d-%b-%Y')} | "
            f"{len(filtered_rows)} kept | {skipped_due} skipped"
        )
        if not filtered_rows:
            raise ValueError(
                "No usable rows match the selected payment dates' Net Due Date window.\n"
                f"Expected Net Due Date between "
                f"{due_start.strftime('%d-%b-%Y')} and {due_end.strftime('%d-%b-%Y')}."
            )
    else:
        skipped_due = 0
    payer_count = len(set(r["payer_name"] for r in filtered_rows))
    progress_cb(
        f"Master: {len(all_rows)} total rows | "
        f"{skipped} skipped (negative) | "
        f"{skipped_due} skipped (Net Due outside target month) | "
        f"{len(filtered_rows)} used | "
        f"{payer_count} unique payers"
    )

    # Determine month/year label (one month before Net Due Date month)
    month_year = get_month_year_label(filtered_rows)
    progress_cb(f"File naming month: {month_year}")

    # Step 4: Archive existing OUTPUT folder contents
    progress_cb("Checking OUTPUT folder for existing files to archive...")
    archive_output_folder(month_year, progress_cb)

    # Step 5: Rebuild Sheet1
    progress_cb("Rebuilding Sheet1...")
    last_row, used_dates = build_sheet1(wb, filtered_rows, dates, daily_rate)
    progress_cb(f"Sheet1 built. {last_row} rows written.")

    # Step 6: Summary table
    progress_cb("Writing summary table...")
    pph_row = build_summary_table(wb, last_row, filtered_rows, n_dcols)
    progress_cb(f"Summary table written at row {pph_row}.")

    # Step 7: Save and recalculate
    progress_cb(f"Saving {SMARTPAYER_FILE}...")
    _safe_save(wb, SMARTPAYER_FILE, progress_cb)
    progress_cb("Recalculating formulas via Excel...")
    ok = recalculate(SMARTPAYER_FILE)
    if not ok:
        progress_cb(
            "Note: Auto-recalculation unavailable. "
            "Open Smartpayer.xlsx in Excel and press Ctrl+Alt+F9 if needed."
        )
    else:
        progress_cb("Recalculation complete.")

    # Step 8: Create new output files
    progress_cb("Creating per-payer output files...")
    created = create_output_files(
        SMARTPAYER_FILE, filtered_rows, dates,
        last_row, n_dcols, daily_rate, month_year, progress_cb
    )
    progress_cb(f"{len(created)} output files created in {OUTPUT_FOLDER}/")

    # Step 9: Archive input file
    input_dir = os.path.abspath(os.path.dirname(input_path))
    default_input_dir = os.path.abspath(INPUT_FOLDER)
    move_input = input_dir == default_input_dir
    progress_cb(
        "Moving input file to BACKUP folder..."
        if move_input else
        "Copying selected input file to BACKUP folder..."
    )
    archived = archive_input_file(input_path, month_year, move=move_input)
    progress_cb(f"Archived as: {os.path.basename(archived)}")

    progress_cb("DONE")
    return created
